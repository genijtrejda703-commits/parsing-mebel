"""Excel export of the approved catalogue (openpyxl, native UTF-8).

Keeps the structural logic of the source data:
    Фабрика -> Коллекция -> Модель -> Категория -> Габариты -> Цена мин/макс
Two shapes are supported:
    mode="product"   one row per product (with its min-max range)
    mode="variation" one row per finish variation (the full price matrix)
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BRASS = "C9A227"
DARK = "1C1A17"
LIGHT = "FAF7F0"

HEAD_PRODUCT = ["Фабрика", "Коллекция", "Модель", "Категория", "Габариты", "Артикул",
                "Вариантов", "Цена мин, €", "Цена макс, €", "Валюта", "Точность, %",
                "Статус", "Документ", "Стр.", "Заметки проверяющего"]
HEAD_VARIATION = ["Фабрика", "Коллекция", "Модель", "Категория", "Габариты", "Артикул",
                  "Отделка", "Цена, €", "Точность, %", "Статус", "Документ", "Стр.",
                  "Заметки проверяющего"]
WIDTHS_PRODUCT = [16, 26, 24, 30, 16, 13, 10, 13, 13, 9, 11, 12, 34, 7, 34]
WIDTHS_VARIATION = [16, 26, 24, 30, 16, 13, 30, 12, 11, 12, 34, 7, 34]

STATUS_RU = {"approved": "Одобрено", "pending": "Ожидает", "rejected": "Отклонено"}

thin = Side(style="thin", color="D8D2C4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _pretty_collection(name):
    import os
    import re
    base = os.path.splitext(str(name or ""))[0]
    parts = [p for p in re.split(r"[_\-\s]+", base)
             if p and p.upper() not in {"PL", "EN", "EUR", "IT"} and not p.isdigit()]
    return " ".join(parts)


def _style_header(ws, headers, widths, row=4):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color=DARK, size=10)
        cell.fill = PatternFill("solid", fgColor=BRASS)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def _title(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(ncols, 6))
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=16, color=DARK)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(ncols, 8))
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(size=9, color="6B6459")
    ws.row_dimensions[1].height = 22


def build_workbook(products, meta):
    mode = meta.get("mode", "product")
    headers = HEAD_VARIATION if mode == "variation" else HEAD_PRODUCT
    widths = WIDTHS_VARIATION if mode == "variation" else WIDTHS_PRODUCT

    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"

    stamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    _title(ws, f"Каталог — {meta.get('factory') or 'все фабрики'}",
           f"Статус: {meta.get('status_label')} · позиций: {len(products)} · "
           f"выгрузка: {stamp} · HOMEART Data Hub", len(headers))
    _style_header(ws, headers, widths)

    def key(p):
        return (str(p.get("factory") or ""), str(p.get("doc_name") or ""),
                str(p.get("model_name") or ""), str(p.get("category") or ""),
                str(p.get("dimension") or ""))
    rows = sorted(products, key=key)

    r = 5
    money_fmt = "#,##0"
    for p in rows:
        coll = _pretty_collection(p.get("doc_name") or p.get("collection"))
        base = [p.get("factory"), coll, p.get("model_name"), p.get("category"),
                p.get("dimension"), p.get("variant_code")]
        if mode == "variation":
            for v in (p.get("variations") or []):
                vals = base + [v.get("finish"), v.get("price"),
                               round((v.get("confidence") or 0) * 100, 1),
                               STATUS_RU.get(p.get("status"), p.get("status")),
                               p.get("doc_name"), (p.get("page") or 0) + 1,
                               p.get("reviewer_notes")]
                _write_row(ws, r, vals, [8], money_fmt)
                r += 1
        else:
            vals = base + [p.get("n_variations"), p.get("price_min"), p.get("price_max"),
                           p.get("currency") or "EUR",
                           round((p.get("confidence") or 0) * 100, 1),
                           STATUS_RU.get(p.get("status"), p.get("status")),
                           p.get("doc_name"), (p.get("page") or 0) + 1,
                           p.get("reviewer_notes")]
            _write_row(ws, r, vals, [8, 9], money_fmt)
            r += 1

    # ---- сводка ----
    ws2 = wb.create_sheet("Сводка")
    _title(ws2, "Сводка по моделям", f"выгрузка: {stamp}", 5)
    _style_header(ws2, ["Модель", "Позиций", "Цена мин, €", "Цена макс, €", "Документ"],
                  [30, 12, 14, 14, 36])
    agg = {}
    for p in rows:
        k = (p.get("model_name"), p.get("doc_name"))
        a = agg.setdefault(k, {"n": 0, "mn": None, "mx": None})
        a["n"] += 1
        for f, v in (("mn", p.get("price_min")), ("mx", p.get("price_max"))):
            if v is None:
                continue
            if a[f] is None:
                a[f] = v
            else:
                a[f] = min(a[f], v) if f == "mn" else max(a[f], v)
    rr = 5
    for (modelname, docname), a in sorted(agg.items(), key=lambda x: -x[1]["n"]):
        _write_row(ws2, rr, [modelname, a["n"], a["mn"], a["mx"], docname], [3, 4], money_fmt)
        rr += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_row(ws, r, vals, money_cols, money_fmt):
    fill = PatternFill("solid", fgColor=LIGHT) if r % 2 == 0 else None
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.font = Font(size=10)
        cell.alignment = Alignment(vertical="center",
                                   wrap_text=c in (4, 7, len(vals)),
                                   horizontal="right" if c in money_cols else "left")
        if c in money_cols:
            cell.number_format = money_fmt
        if fill:
            cell.fill = fill


HEAD_POS = ["Фабрика", "Модель", "Категории", "Вариантов", "Цена мин, €",
            "Цена макс, €", "Валюта", "Точность, %", "Статус", "Документов",
            "Страниц", "Заметки проверяющего"]
WIDTHS_POS = [16, 28, 34, 11, 13, 13, 9, 11, 12, 11, 9, 34]
HEAD_PRICE = ["Фабрика", "Модель", "Артикул", "Габариты", "Отделка", "Цена, €",
              "Валюта", "Точность, %", "Статус", "Документ", "Стр."]
WIDTHS_PRICE = [16, 26, 14, 18, 30, 12, 9, 11, 12, 34, 7]


def build_position_workbook(positions, variants, meta):
    """Position-first catalogue: sheets «Позиции», «Цены», «Сводка»."""
    stamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    wb = Workbook()

    # ---- Позиции ----
    ws = wb.active
    ws.title = "Позиции"
    _title(ws, f"Позиции — {meta.get('factory') or 'все фабрики'}",
           f"Статус: {meta.get('status_label')} · позиций: {len(positions)} · "
           f"вариантов-цен: {len(variants)} · выгрузка: {stamp} · HOMEART Data Hub",
           len(HEAD_POS))
    _style_header(ws, HEAD_POS, WIDTHS_POS)
    r = 5
    for p in sorted(positions, key=lambda x: (str(x.get("factory") or ""),
                                              str(x.get("name") or ""))):
        cats = ", ".join((p.get("categories") or [])[:6])
        vals = [p.get("factory"), p.get("name"), cats, p.get("n_variants"),
                p.get("price_min"), p.get("price_max"), p.get("currency") or "EUR",
                round((p.get("avg_confidence") or 0) * 100, 1),
                STATUS_RU.get(p.get("status"), p.get("status")),
                p.get("n_docs"), p.get("n_pages"), p.get("reviewer_notes")]
        _write_row(ws, r, vals, [5, 6], "#,##0")
        r += 1

    # ---- Цены ----
    ws2 = wb.create_sheet("Цены")
    _title(ws2, "Варианты и цены", f"строк: {len(variants)} · выгрузка: {stamp}",
           len(HEAD_PRICE))
    _style_header(ws2, HEAD_PRICE, WIDTHS_PRICE)
    r = 5
    for v in sorted(variants, key=lambda x: (str(x.get("position_name") or ""),
                                             str(x.get("variant_code") or ""),
                                             str(x.get("finish") or ""))):
        vals = [v.get("factory"), v.get("position_name"), v.get("variant_code"),
                v.get("dimension"), v.get("finish"), v.get("price"),
                v.get("currency") or "EUR",
                round((v.get("confidence") or 0) * 100, 1),
                STATUS_RU.get(v.get("status"), v.get("status")),
                v.get("doc_name"), (v.get("page") or 0) + 1]
        _write_row(ws2, r, vals, [6], "#,##0")
        r += 1

    # ---- Сводка ----
    ws3 = wb.create_sheet("Сводка")
    _title(ws3, "Сводка по документам", f"выгрузка: {stamp}", 5)
    _style_header(ws3, ["Документ", "Позиций", "Вариантов-цен",
                        "Цена мин, €", "Цена макс, €"], [40, 12, 14, 14, 14])
    by_doc = {}
    for v in variants:
        d = by_doc.setdefault(v.get("doc_name") or "—",
                              {"pos": set(), "n": 0, "mn": None, "mx": None})
        d["pos"].add(v.get("position_id"))
        d["n"] += 1
        pr = v.get("price")
        if pr is not None:
            d["mn"] = pr if d["mn"] is None else min(d["mn"], pr)
            d["mx"] = pr if d["mx"] is None else max(d["mx"], pr)
    rr = 5
    for docname, d in sorted(by_doc.items(), key=lambda x: -x[1]["n"]):
        _write_row(ws3, rr, [docname, len(d["pos"]), d["n"], d["mn"], d["mx"]],
                   [4, 5], "#,##0")
        rr += 1
    _write_row(ws3, rr, ["ИТОГО", len(positions), len(variants), None, None],
               [4, 5], "#,##0")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


HEAD_INV = ["Папка", "Файл", "Тип документа", "Год", "Валюта", "Страниц",
            "Актуальный", "Разобран", "Позиций", "Вариантов-цен", "Замена",
            "Заголовок (из содержимого)"]
WIDTHS_INV = [26, 40, 18, 8, 9, 9, 12, 11, 10, 13, 22, 40]


def build_inventory_workbook(records, docs):
    """Sheet «Инвентаризация» (file classification) + «Покрытие» (coverage)."""
    stamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    wb = Workbook()
    ws = wb.active
    ws.title = "Инвентаризация"
    cur = sum(1 for r in records if r.get("is_current_listino"))
    parsed = sum(1 for r in records if r.get("ingested"))
    _title(ws, "Инвентаризация файлов фабрики",
           f"файлов: {len(records)} · актуальных прайс-листов: {cur} · "
           f"разобрано: {parsed} · выгрузка: {stamp}", len(HEAD_INV))
    _style_header(ws, HEAD_INV, WIDTHS_INV)
    r = 5
    for rec in sorted(records, key=lambda x: (str(x.get("doc_type") or ""),
                                              -(x.get("year") or 0),
                                              str(x.get("name") or ""))):
        vals = [rec.get("folder") or "/", rec.get("name"), rec.get("doc_type"),
                rec.get("year"), rec.get("currency"), rec.get("pages"),
                "да" if rec.get("is_current_listino") else "—",
                "да" if rec.get("ingested") else "—",
                rec.get("positions"), rec.get("variant_prices"),
                rec.get("superseded_by") or "—", rec.get("sample_title")]
        _write_row(ws, r, vals, [], "#,##0")
        r += 1

    ws2 = wb.create_sheet("Покрытие")
    _title(ws2, "Покрытие по документам",
           "страниц всего / с матрицами / разобрано / пропущено · позиций · "
           f"вариантов-цен · выгрузка: {stamp}", 7)
    _style_header(ws2, ["Документ", "Страниц", "С матрицами", "Разобрано",
                        "Пропущено", "Позиций", "Вариантов-цен"],
                  [40, 10, 13, 12, 12, 10, 14])
    rr = 5
    T = [0, 0, 0, 0, 0, 0]
    for d in sorted(docs, key=lambda x: -(x.get("variant_prices") or 0)):
        c = d.get("coverage") or {}
        row = [c.get("pages_total") or d.get("pages") or 0, c.get("pages_with_matrix") or 0,
               c.get("pages_parsed") or 0, c.get("pages_skipped") or 0,
               d.get("positions") or 0, d.get("variant_prices") or 0]
        _write_row(ws2, rr, [d.get("name")] + row, [], "#,##0")
        for i in range(6):
            T[i] += row[i]
        rr += 1
    _write_row(ws2, rr, ["ИТОГО"] + T, [], "#,##0")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
