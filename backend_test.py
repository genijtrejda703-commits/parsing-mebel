#!/usr/bin/env python3
"""Backend API test suite for HOMEART Data Hub - Slice 3 features.

Tests the 4 new backend features:
1. Excel export (GET /api/export)
2. Hybrid RU/EN search (POST /api/search with lexicon)
3. Anomaly review lane (GET /api/anomalies)
4. Product illustration crops (GET /api/product-photo)
5. Regression smoke tests
"""
import io
import os
import sys
import time
import requests
from PIL import Image

# Backend URL from environment
BASE_URL = os.environ.get("NEXT_PUBLIC_BASE_URL", "https://saas-ingestion.preview.emergentagent.com")
API_BASE = f"{BASE_URL}/api"

# Test configuration
TIMEOUT = 180  # Allow generous timeout for large exports


def test_excel_export():
    """Test A: Excel export - GET /api/export"""
    print("\n" + "="*80)
    print("TEST A: Excel export (GET /api/export)")
    print("="*80)
    
    results = {"passed": 0, "failed": 0, "tests": []}
    
    # Test A1: mode=product, status=approved (60 rows expected)
    try:
        print("\n[A1] Testing mode=product, status=approved...")
        url = f"{API_BASE}/export?status=approved&mode=product"
        resp = requests.get(url, timeout=TIMEOUT)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        # Check Content-Type
        ct = resp.headers.get("Content-Type", "")
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in ct, \
            f"Wrong Content-Type: {ct}"
        
        # Check X-Rows header
        x_rows = resp.headers.get("X-Rows")
        assert x_rows is not None, "Missing X-Rows header"
        rows = int(x_rows)
        print(f"   X-Rows: {rows} (expected ~60)")
        assert rows >= 50 and rows <= 70, f"Expected ~60 rows, got {rows}"
        
        # Check body starts with ZIP magic bytes 'PK'
        body = resp.content
        assert len(body) > 5000, f"Body too small: {len(body)} bytes"
        assert body[:2] == b'PK', f"Missing ZIP magic bytes, got: {body[:4].hex()}"
        
        print(f"   ✅ PASSED - mode=product: {len(body)} bytes, {rows} rows, valid xlsx")
        results["passed"] += 1
        results["tests"].append(("A1: mode=product status=approved", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("A1: mode=product status=approved", False, str(e)))
    
    # Test A2: mode=variation (should be LARGER than mode=product)
    try:
        print("\n[A2] Testing mode=variation, status=approved...")
        url = f"{API_BASE}/export?status=approved&mode=variation"
        resp = requests.get(url, timeout=TIMEOUT)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        body_var = resp.content
        x_rows_var = int(resp.headers.get("X-Rows", "0"))
        
        print(f"   mode=variation: {len(body_var)} bytes, {x_rows_var} rows")
        print(f"   mode=product:   {len(body)} bytes, {rows} rows")
        
        # Variation mode should have MORE rows (one per finish)
        assert x_rows_var > rows, \
            f"mode=variation should have more rows than mode=product ({x_rows_var} vs {rows})"
        
        # Body should be larger
        assert len(body_var) > len(body), \
            f"mode=variation body should be larger ({len(body_var)} vs {len(body)})"
        
        print(f"   ✅ PASSED - mode=variation is larger: {x_rows_var} rows vs {rows} rows")
        results["passed"] += 1
        results["tests"].append(("A2: mode=variation larger than product", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("A2: mode=variation larger than product", False, str(e)))
    
    # Test A3: status=all (should be ~20,499 rows, allow generous timeout)
    try:
        print("\n[A3] Testing status=all (large export, ~20,499 rows)...")
        url = f"{API_BASE}/export?status=all&mode=product"
        resp = requests.get(url, timeout=TIMEOUT)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        x_rows_all = int(resp.headers.get("X-Rows", "0"))
        print(f"   X-Rows: {x_rows_all} (expected ~20,499)")
        
        assert x_rows_all >= 20000 and x_rows_all <= 21000, \
            f"Expected ~20,499 rows, got {x_rows_all}"
        
        print(f"   ✅ PASSED - status=all: {x_rows_all} rows, {len(resp.content)} bytes")
        results["passed"] += 1
        results["tests"].append(("A3: status=all large export", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("A3: status=all large export", False, str(e)))
    
    # Test A4: status=rejected (0 rows, but should still return valid xlsx)
    try:
        print("\n[A4] Testing status=rejected (0 rows, should still be valid xlsx)...")
        url = f"{API_BASE}/export?status=rejected&mode=product"
        resp = requests.get(url, timeout=TIMEOUT)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        x_rows_rej = int(resp.headers.get("X-Rows", "0"))
        print(f"   X-Rows: {x_rows_rej} (expected 0)")
        
        # Should still be valid xlsx with headers
        assert resp.content[:2] == b'PK', "Should still return valid xlsx"
        assert len(resp.content) > 3000, "Should have headers and structure"
        
        print(f"   ✅ PASSED - status=rejected: valid xlsx with {x_rows_rej} rows")
        results["passed"] += 1
        results["tests"].append(("A4: status=rejected empty but valid", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("A4: status=rejected empty but valid", False, str(e)))
    
    # Test A5: Open with openpyxl and verify structure
    try:
        print("\n[A5] Testing xlsx structure with openpyxl...")
        from openpyxl import load_workbook
        
        # Use the approved product export from A1
        wb = load_workbook(io.BytesIO(body))
        sheet_names = wb.sheetnames
        
        print(f"   Sheet names: {sheet_names}")
        assert sheet_names == ['Каталог', 'Сводка'], \
            f"Expected ['Каталог', 'Сводка'], got {sheet_names}"
        
        # Check row 4 of 'Каталог' has Cyrillic headers
        ws = wb['Каталог']
        row4 = [cell.value for cell in ws[4]]
        print(f"   Row 4 headers: {row4[:6]}")
        
        expected_start = ['Фабрика', 'Коллекция', 'Модель', 'Категория', 'Габариты', 'Артикул']
        assert row4[:6] == expected_start, \
            f"Expected {expected_start}, got {row4[:6]}"
        
        # Check freeze panes are set
        assert ws.freeze_panes is not None, "Freeze panes not set"
        print(f"   Freeze panes: {ws.freeze_panes}")
        
        # Verify Cyrillic is not mojibake (check a few cells)
        assert 'Фабрика' in str(row4[0]), "Cyrillic appears to be mojibake"
        
        print(f"   ✅ PASSED - xlsx structure correct: sheets, headers, freeze panes")
        results["passed"] += 1
        results["tests"].append(("A5: xlsx structure with openpyxl", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("A5: xlsx structure with openpyxl", False, str(e)))
    
    print(f"\n{'='*80}")
    print(f"TEST A SUMMARY: {results['passed']}/5 passed, {results['failed']}/5 failed")
    print(f"{'='*80}")
    
    return results


def test_hybrid_search():
    """Test B: Hybrid RU/EN search with lexicon"""
    print("\n" + "="*80)
    print("TEST B: Hybrid RU/EN search (POST /api/search with lexicon)")
    print("="*80)
    
    results = {"passed": 0, "failed": 0, "tests": []}
    
    # Test B0: Unit-check lexicon directly
    try:
        print("\n[B0] Testing lexicon directly...")
        sys.path.insert(0, '/app/backend')
        from lexicon import translate_query
        
        # Test 'обеденный стол' -> should give 'dining' and 'table', NOT 'worktop'
        terms1, matched1 = translate_query('обеденный стол')
        print(f"   translate_query('обеденный стол') -> {terms1}, matched: {matched1}")
        assert 'dining' in terms1, f"Expected 'dining' in {terms1}"
        assert 'table' in terms1, f"Expected 'table' in {terms1}"
        assert 'worktop' not in terms1, f"Should NOT have 'worktop' in {terms1} (stem bug)"
        
        # Test 'кровать' -> should give 'bed'
        terms2, matched2 = translate_query('кровать')
        print(f"   translate_query('кровать') -> {terms2}, matched: {matched2}")
        assert 'bed' in terms2, f"Expected 'bed' in {terms2}"
        
        # Test 'столешница' -> SHOULD map to 'worktop'
        terms3, matched3 = translate_query('столешница')
        print(f"   translate_query('столешница') -> {terms3}, matched: {matched3}")
        assert 'worktop' in terms3 or 'top' in terms3, \
            f"Expected 'worktop' or 'top' in {terms3}"
        
        print(f"   ✅ PASSED - Lexicon working correctly")
        results["passed"] += 1
        results["tests"].append(("B0: lexicon unit test", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("B0: lexicon unit test", False, str(e)))
    
    # Test B1: 'обеденный стол' -> table results
    try:
        print("\n[B1] Testing Russian query 'обеденный стол' (dining table)...")
        url = f"{API_BASE}/search"
        payload = {"q": "обеденный стол", "top_k": 8}
        resp = requests.post(url, json=payload, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        print(f"   Response keys: {data.keys()}")
        
        # Check response structure
        assert "translated_from" in data, "Missing 'translated_from'"
        assert "lexical_terms" in data, "Missing 'lexical_terms'"
        assert "lexical_hits" in data, "Missing 'lexical_hits'"
        assert "results" in data, "Missing 'results'"
        
        print(f"   translated_from: {data['translated_from']}")
        print(f"   lexical_terms: {data['lexical_terms']}")
        print(f"   lexical_hits: {data['lexical_hits']}")
        
        # Check translated_from contains Russian words
        assert 'обеденный' in data['translated_from'] or 'стол' in data['translated_from'], \
            f"Expected Russian words in translated_from: {data['translated_from']}"
        
        # Check lexical_terms contains English translations
        lex_str = ' '.join(data['lexical_terms'])
        assert 'dining' in lex_str or 'table' in lex_str, \
            f"Expected 'dining' or 'table' in lexical_terms: {data['lexical_terms']}"
        
        # Check lexical_hits > 0
        assert data['lexical_hits'] > 0, f"Expected lexical_hits > 0, got {data['lexical_hits']}"
        
        # Check results contain table-related items
        results_list = data['results']
        print(f"   Results count: {len(results_list)}")
        
        table_keywords = ['TABLE', 'DINE', 'SNACK', 'VICINO', 'MATEO', 'DINING']
        found_table = False
        for r in results_list[:5]:
            model = r.get('model_name', '').upper()
            cat = r.get('category', '').upper()
            print(f"     - {r.get('model_name')} / {r.get('category')} (match={r.get('match'):.3f})")
            if any(kw in model or kw in cat for kw in table_keywords):
                found_table = True
        
        assert found_table, "Expected at least one table-related result"
        
        # Check all results have required fields
        for r in results_list:
            assert 'n_variations' in r, "Missing n_variations"
            assert r['n_variations'] >= 2, f"n_variations should be >= 2, got {r['n_variations']}"
            assert 'category' in r and r['category'], "Missing or empty category"
            assert 'match' in r, "Missing match score"
            assert 'score' in r, "Missing score"
            assert 'lex' in r, "Missing lex"
            assert 'price_min' in r, "Missing price_min"
            assert 'price_max' in r, "Missing price_max"
            assert 'doc_id' in r, "Missing doc_id"
            assert 'page' in r, "Missing page"
        
        # Check results are sorted by DESCENDING match
        matches = [r['match'] for r in results_list]
        assert matches == sorted(matches, reverse=True), \
            f"Results not sorted by descending match: {matches}"
        
        # Check no duplicate (model_name, category) pairs
        seen = set()
        for r in results_list:
            key = (r.get('model_name'), r.get('category'))
            assert key not in seen, f"Duplicate (model_name, category): {key}"
            seen.add(key)
        
        print(f"   ✅ PASSED - 'обеденный стол' returns table results with hybrid search")
        results["passed"] += 1
        results["tests"].append(("B1: обеденный стол query", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("B1: обеденный стол query", False, str(e)))
    
    # Test B2: 'кровать' -> bed results
    try:
        print("\n[B2] Testing Russian query 'кровать' (bed)...")
        url = f"{API_BASE}/search"
        payload = {"q": "кровать", "top_k": 8}
        resp = requests.post(url, json=payload, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        results_list = data['results']
        
        bed_keywords = ['BED', 'КРОВАТЬ', 'HEADBOARD', 'MATTRESS']
        found_bed = False
        for r in results_list[:5]:
            model = r.get('model_name', '').upper()
            cat = r.get('category', '').upper()
            print(f"     - {r.get('model_name')} / {r.get('category')}")
            if any(kw in model or kw in cat for kw in bed_keywords):
                found_bed = True
        
        assert found_bed, "Expected at least one bed-related result"
        
        print(f"   ✅ PASSED - 'кровать' returns bed results")
        results["passed"] += 1
        results["tests"].append(("B2: кровать query", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("B2: кровать query", False, str(e)))
    
    # Test B3: 'диван из кожи' -> sofa results (regression test)
    try:
        print("\n[B3] Testing Russian query 'диван из кожи' (leather sofa) - regression...")
        url = f"{API_BASE}/search"
        payload = {"q": "диван из кожи", "top_k": 8}
        resp = requests.post(url, json=payload, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        results_list = data['results']
        
        sofa_keywords = ['GREGOR', 'LUCIO', 'CLEO', 'AUGUSTO', 'OCTAVE', 'SOFA']
        found_sofa = False
        for r in results_list[:5]:
            model = r.get('model_name', '').upper()
            cat = r.get('category', '').upper()
            print(f"     - {r.get('model_name')} / {r.get('category')}")
            if any(kw in model or kw in cat for kw in sofa_keywords):
                found_sofa = True
        
        assert found_sofa, "Expected at least one sofa-related result (regression)"
        
        print(f"   ✅ PASSED - 'диван из кожи' still returns sofa results")
        results["passed"] += 1
        results["tests"].append(("B3: диван из кожи regression", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("B3: диван из кожи regression", False, str(e)))
    
    # Test B4: Empty query -> empty results, not 500
    try:
        print("\n[B4] Testing empty query (should return empty results, not 500)...")
        url = f"{API_BASE}/search"
        payload = {"q": "   ", "top_k": 8}
        resp = requests.post(url, json=payload, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        assert data['results'] == [], f"Expected empty results, got {len(data['results'])}"
        assert data['mode'] == 'text', f"Expected mode='text', got {data['mode']}"
        
        print(f"   ✅ PASSED - Empty query returns empty results with 200")
        results["passed"] += 1
        results["tests"].append(("B4: empty query", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("B4: empty query", False, str(e)))
    
    # Test B5: Image mode still works
    try:
        print("\n[B5] Testing image mode (upload small JPEG)...")
        
        # Create a small test image with PIL
        img = Image.new('RGB', (400, 300), color='red')
        buf = io.BytesIO()
        img.save(buf, format='JPEG')
        buf.seek(0)
        
        url = f"{API_BASE}/search"
        files = {'file': ('test.jpg', buf, 'image/jpeg')}
        data = {'top_k': '8'}
        resp = requests.post(url, files=files, data=data, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        result = resp.json()
        assert result['mode'] == 'image', f"Expected mode='image', got {result['mode']}"
        assert len(result['results']) == 8, f"Expected 8 results, got {len(result['results'])}"
        
        # Check scores are in expected CLIP image-text band (0.1-0.45)
        scores = [r['score'] for r in result['results']]
        print(f"   Image scores: {scores}")
        assert all(0.1 <= s <= 0.5 for s in scores), \
            f"Image scores outside expected range: {scores}"
        
        print(f"   ✅ PASSED - Image mode returns 8 results with scores in expected range")
        results["passed"] += 1
        results["tests"].append(("B5: image mode", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("B5: image mode", False, str(e)))
    
    print(f"\n{'='*80}")
    print(f"TEST B SUMMARY: {results['passed']}/6 passed, {results['failed']}/6 failed")
    print(f"{'='*80}")
    
    return results


def test_anomaly_review():
    """Test C: Anomaly review lane - GET /api/anomalies"""
    print("\n" + "="*80)
    print("TEST C: Anomaly review lane (GET /api/anomalies)")
    print("="*80)
    
    results = {"passed": 0, "failed": 0, "tests": []}
    
    # Test C1: Basic query with limit=5
    try:
        print("\n[C1] Testing basic query with limit=5...")
        url = f"{API_BASE}/anomalies?limit=5"
        resp = requests.get(url, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        print(f"   Response keys: {data.keys()}")
        
        assert 'items' in data, "Missing 'items'"
        assert 'total' in data, "Missing 'total'"
        assert 'reasons' in data, "Missing 'reasons'"
        
        total = data['total']
        print(f"   Total anomalies: {total} (expected ~23,042)")
        assert total >= 22000 and total <= 24000, \
            f"Expected ~23,042 anomalies, got {total}"
        
        # Check reasons facets
        reasons = data['reasons']
        print(f"   Reasons count: {len(reasons)}")
        assert len(reasons) >= 5, f"Expected at least 5 reason categories, got {len(reasons)}"
        
        # Check for expected Russian reason categories
        reason_texts = [r['reason'] for r in reasons]
        print(f"   Reason categories: {reason_texts}")
        
        expected_reasons = [
            'Колонтитул страницы (верхние/нижние 5%)',
            'Нет заголовка столбца сверху'
        ]
        for exp in expected_reasons:
            assert any(exp in r for r in reason_texts), \
                f"Expected reason '{exp}' not found in {reason_texts}"
        
        # Check items structure
        items = data['items']
        assert len(items) == 5, f"Expected 5 items, got {len(items)}"
        
        # Check default sort is confidence DESCENDING
        confidences = [item['confidence'] for item in items]
        print(f"   Confidences: {confidences}")
        assert confidences[0] >= confidences[-1], \
            f"Items not sorted by descending confidence: {confidences}"
        
        # Check required fields in each item
        for item in items:
            assert 'text' in item, "Missing 'text'"
            assert 'confidence' in item, "Missing 'confidence'"
            assert 'row_peers' in item, "Missing 'row_peers'"
            assert 'col_peers' in item, "Missing 'col_peers'"
            assert 'above_text' in item, "Missing 'above_text'"
            assert 'left_text' in item, "Missing 'left_text'"
            assert 'reason' in item, "Missing 'reason'"
            assert 'doc_name' in item, "Missing 'doc_name'"
            assert 'page' in item, "Missing 'page'"
            assert '_id' not in item, "Should not expose Mongo _id"
        
        print(f"   ✅ PASSED - Basic query returns {total} anomalies with correct structure")
        results["passed"] += 1
        results["tests"].append(("C1: basic anomalies query", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("C1: basic anomalies query", False, str(e)))
    
    # Test C2: Filter by reason
    try:
        print("\n[C2] Testing filter by reason...")
        
        # Pick the first reason from the facets
        first_reason = reasons[0]['reason']
        expected_count = reasons[0]['n']
        
        print(f"   Filtering by reason: '{first_reason}' (expected {expected_count} items)")
        
        url = f"{API_BASE}/anomalies?reason={first_reason}&limit=50"
        resp = requests.get(url, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        assert data['total'] == expected_count, \
            f"Expected {expected_count} items, got {data['total']}"
        
        # Check all returned items have exactly that reason
        for item in data['items']:
            assert item['reason'] == first_reason, \
                f"Expected reason '{first_reason}', got '{item['reason']}'"
        
        print(f"   ✅ PASSED - Reason filter returns {data['total']} items, all with correct reason")
        results["passed"] += 1
        results["tests"].append(("C2: filter by reason", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("C2: filter by reason", False, str(e)))
    
    # Test C3: Pagination
    try:
        print("\n[C3] Testing pagination (skip parameter)...")
        
        url1 = f"{API_BASE}/anomalies?limit=5&skip=0"
        resp1 = requests.get(url1, timeout=30)
        data1 = resp1.json()
        first_item_page0 = data1['items'][0]['text'] if data1['items'] else None
        
        url2 = f"{API_BASE}/anomalies?limit=5&skip=5"
        resp2 = requests.get(url2, timeout=30)
        data2 = resp2.json()
        first_item_page1 = data2['items'][0]['text'] if data2['items'] else None
        
        print(f"   First item (skip=0): {first_item_page0[:50] if first_item_page0 else 'None'}...")
        print(f"   First item (skip=5): {first_item_page1[:50] if first_item_page1 else 'None'}...")
        
        assert first_item_page0 != first_item_page1, \
            "Pagination not working - same first item on different pages"
        
        print(f"   ✅ PASSED - Pagination returns different items")
        results["passed"] += 1
        results["tests"].append(("C3: pagination", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("C3: pagination", False, str(e)))
    
    # Test C4: Text search filter (q parameter)
    try:
        print("\n[C4] Testing text search filter (q parameter)...")
        
        # Pick a substring from a returned model_name or text
        if items and items[0].get('text'):
            search_term = items[0]['text'].split()[0] if items[0]['text'].split() else 'table'
        else:
            search_term = 'table'
        
        print(f"   Searching for: '{search_term}'")
        
        url = f"{API_BASE}/anomalies?q={search_term}&limit=20"
        resp = requests.get(url, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        print(f"   Found {data['total']} items matching '{search_term}'")
        
        # Check that results contain the search term (case-insensitive)
        found = False
        for item in data['items'][:5]:
            text_fields = [
                item.get('text', ''),
                item.get('model_name', ''),
                item.get('category', ''),
                item.get('above_text', ''),
                item.get('left_text', '')
            ]
            if any(search_term.lower() in str(f).lower() for f in text_fields):
                found = True
                break
        
        assert found or data['total'] == 0, \
            f"Search results don't contain '{search_term}'"
        
        print(f"   ✅ PASSED - Text search returns filtered results")
        results["passed"] += 1
        results["tests"].append(("C4: text search filter", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("C4: text search filter", False, str(e)))
    
    print(f"\n{'='*80}")
    print(f"TEST C SUMMARY: {results['passed']}/4 passed, {results['failed']}/4 failed")
    print(f"{'='*80}")
    
    return results


def test_product_photos():
    """Test D: Product illustration crops - GET /api/product-photo"""
    print("\n" + "="*80)
    print("TEST D: Product illustration crops (GET /api/product-photo)")
    print("="*80)
    
    results = {"passed": 0, "failed": 0, "tests": []}
    
    # Test D1: Find a product with a photo
    try:
        print("\n[D1] Finding a product with a photo...")
        
        # Query products with high confidence and multiple variations
        url = f"{API_BASE}/products?limit=50&min_conf=0.8&min_var=3&sort=best"
        resp = requests.get(url, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        products = data['items']
        
        # Find a product with a photo field
        product_with_photo = None
        for p in products:
            if 'photo' in p and p['photo']:
                product_with_photo = p
                break
        
        if not product_with_photo:
            # Try querying more products
            print("   No photo in first 50, trying more...")
            url = f"{API_BASE}/products?limit=100&sort=best"
            resp = requests.get(url, timeout=30)
            data = resp.json()
            for p in data['items']:
                if 'photo' in p and p['photo']:
                    product_with_photo = p
                    break
        
        assert product_with_photo is not None, \
            "Could not find any product with a photo field"
        
        product_id = product_with_photo['id']
        photo = product_with_photo['photo']
        
        print(f"   Found product with photo: {product_id}")
        print(f"   Photo: page={photo.get('page')}, bbox={photo.get('bbox')}")
        
        # Test D2: GET /api/product-photo with valid product_id
        print("\n[D2] Testing GET /api/product-photo with valid product_id...")
        
        url = f"{API_BASE}/product-photo?product_id={product_id}"
        resp = requests.get(url, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        # Check Content-Type
        ct = resp.headers.get('Content-Type', '')
        assert ct == 'image/png', f"Expected 'image/png', got '{ct}'"
        
        # Check PNG magic bytes
        body = resp.content
        assert body[:4] == b'\x89PNG', \
            f"Expected PNG magic bytes, got {body[:4].hex()}"
        
        # Check size > 1 KB
        assert len(body) > 1024, f"Image too small: {len(body)} bytes"
        
        print(f"   ✅ PASSED - Returns valid PNG image ({len(body)} bytes)")
        results["passed"] += 1
        results["tests"].append(("D2: product-photo valid id", True, None))
        
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("D2: product-photo valid id", False, str(e)))
    
    # Test D3: GET /api/product-photo with invalid product_id
    try:
        print("\n[D3] Testing GET /api/product-photo with invalid product_id...")
        
        url = f"{API_BASE}/product-photo?product_id=doesnotexist"
        resp = requests.get(url, timeout=30)
        
        # Should return 4xx error, not 500
        assert resp.status_code >= 400 and resp.status_code < 500, \
            f"Expected 4xx error, got {resp.status_code}"
        
        print(f"   ✅ PASSED - Returns {resp.status_code} error for invalid product_id")
        results["passed"] += 1
        results["tests"].append(("D3: product-photo invalid id", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("D3: product-photo invalid id", False, str(e)))
    
    # Test D4: Count products with photos
    try:
        print("\n[D4] Counting products with photos...")
        
        # Use MongoDB directly (read-only)
        from pymongo import MongoClient
        client = MongoClient(os.environ.get('MONGO_URL', 'mongodb://localhost:27017'))
        db = client[os.environ.get('DB_NAME', 'your_database_name')]
        
        count = db.products.count_documents({'photo': {'$exists': True}})
        
        print(f"   Products with photos: {count} (expected ~14,356)")
        assert count >= 14000 and count <= 15000, \
            f"Expected ~14,356 products with photos, got {count}"
        
        print(f"   ✅ PASSED - {count} products have photos")
        results["passed"] += 1
        results["tests"].append(("D4: count products with photos", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("D4: count products with photos", False, str(e)))
    
    print(f"\n{'='*80}")
    print(f"TEST D SUMMARY: {results['passed']}/3 passed, {results['failed']}/3 failed")
    print(f"{'='*80}")
    
    return results


def test_regression_smoke():
    """Test E: Regression smoke tests (quick, 1 call each)"""
    print("\n" + "="*80)
    print("TEST E: Regression smoke tests")
    print("="*80)
    
    results = {"passed": 0, "failed": 0, "tests": []}
    
    # Test E1: GET /api/stats
    try:
        print("\n[E1] Testing GET /api/stats...")
        url = f"{API_BASE}/stats"
        resp = requests.get(url, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        print(f"   Stats: products={data.get('products')}, approved={data.get('approved')}, "
              f"embeddings={data.get('embeddings')}, documents={data.get('documents')}")
        
        assert data['products'] == 20499, f"Expected 20499 products, got {data['products']}"
        assert data['approved'] == 60, f"Expected 60 approved, got {data['approved']}"
        assert data['embeddings'] == 16309, f"Expected 16309 embeddings, got {data['embeddings']}"
        assert data['documents'] == 5, f"Expected 5 documents, got {data['documents']}"
        
        print(f"   ✅ PASSED - Stats correct")
        results["passed"] += 1
        results["tests"].append(("E1: stats", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("E1: stats", False, str(e)))
    
    # Test E2: GET /api/products
    try:
        print("\n[E2] Testing GET /api/products?limit=3&sort=best&min_var=2...")
        url = f"{API_BASE}/products?limit=3&sort=best&min_var=2"
        resp = requests.get(url, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        assert len(data['items']) == 3, f"Expected 3 items, got {len(data['items'])}"
        
        print(f"   ✅ PASSED - Products query returns 3 items")
        results["passed"] += 1
        results["tests"].append(("E2: products query", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("E2: products query", False, str(e)))
    
    # Test E3: GET /api/tasks
    try:
        print("\n[E3] Testing GET /api/tasks...")
        url = f"{API_BASE}/tasks"
        resp = requests.get(url, timeout=30)
        
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        
        data = resp.json()
        assert 'items' in data, "Missing 'items'"
        
        # Check task types include expected types
        task_types = [t.get('type') for t in data['items']]
        print(f"   Task types: {set(task_types)}")
        
        expected_types = ['scan', 'ingest', 'embed', 'photos', 'anomalies']
        found_types = [t for t in expected_types if t in task_types]
        
        print(f"   Found task types: {found_types}")
        assert len(found_types) >= 3, \
            f"Expected at least 3 task types from {expected_types}, found {found_types}"
        
        print(f"   ✅ PASSED - Tasks list includes expected types")
        results["passed"] += 1
        results["tests"].append(("E3: tasks list", True, None))
    except Exception as e:
        print(f"   ❌ FAILED - {e}")
        results["failed"] += 1
        results["tests"].append(("E3: tasks list", False, str(e)))
    
    print(f"\n{'='*80}")
    print(f"TEST E SUMMARY: {results['passed']}/3 passed, {results['failed']}/3 failed")
    print(f"{'='*80}")
    
    return results


def main():
    """Run all backend tests"""
    print("\n" + "="*80)
    print("HOMEART DATA HUB - BACKEND TEST SUITE (Slice 3)")
    print("="*80)
    print(f"Backend URL: {API_BASE}")
    print(f"Timeout: {TIMEOUT}s")
    print("="*80)
    
    all_results = {
        "A": test_excel_export(),
        "B": test_hybrid_search(),
        "C": test_anomaly_review(),
        "D": test_product_photos(),
        "E": test_regression_smoke(),
    }
    
    # Summary
    print("\n" + "="*80)
    print("FINAL SUMMARY")
    print("="*80)
    
    total_passed = sum(r["passed"] for r in all_results.values())
    total_failed = sum(r["failed"] for r in all_results.values())
    total_tests = total_passed + total_failed
    
    for section, result in all_results.items():
        status = "✅ PASS" if result["failed"] == 0 else "❌ FAIL"
        print(f"  {section}: {result['passed']}/{result['passed'] + result['failed']} passed {status}")
    
    print(f"\n{'='*80}")
    print(f"OVERALL: {total_passed}/{total_tests} tests passed ({100*total_passed//total_tests}%)")
    print(f"{'='*80}")
    
    # Detailed failures
    if total_failed > 0:
        print("\n" + "="*80)
        print("FAILED TESTS DETAILS")
        print("="*80)
        for section, result in all_results.items():
            for test_name, passed, error in result["tests"]:
                if not passed:
                    print(f"\n❌ {section} - {test_name}")
                    print(f"   Error: {error}")
    
    return 0 if total_failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
